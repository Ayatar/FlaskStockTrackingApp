from flask import Flask, render_template, redirect, url_for, flash, request, send_file, abort
from config import Config
from models import db, Category, Product, StockMovement
from forms import ProductForm, CategoryForm, StockMovementForm, ReportForm
from datetime import datetime, timedelta
from sqlalchemy import func
import pandas as pd
from openpyxl import Workbook
from io import BytesIO
from flask_wtf.csrf import CSRFProtect, generate_csrf

app = Flask(__name__)
app.config.from_object(Config)

csrf = CSRFProtect(app)

db.init_app(app)

with app.app_context():
    db.create_all()

    #if first category don't exist add a example category
    if Category.query.count() == 0:
        categories = [
            Category(name='Electronics', description='Electronics Products'),
            Category(name='Nutrient', description='Nutrient Products'),
            Category(name='Textile', description='Textile Products')
        ]
        db.session.add_all(categories)
        db.session.commit()


@app.context_processor
def inject_csrf_token():
    """Expose CSRF token helper to all templates."""
    return dict(csrf_token=generate_csrf)


@app.after_request
def add_security_headers(response):
    """Apply basic security headers to every response."""
    csp = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; "
        "style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; "
        "img-src 'self' data:; "
        "font-src 'self' https://cdnjs.cloudflare.com; "
        "connect-src 'self'; "
        "frame-ancestors 'self';"
    )

    response.headers.setdefault('Content-Security-Policy', csp)
    response.headers.setdefault('X-Content-Type-Options', 'nosniff')
    response.headers.setdefault('X-Frame-Options', 'SAMEORIGIN')
    response.headers.setdefault('X-XSS-Protection', '1; mode=block')
    response.headers.setdefault('Referrer-Policy', 'strict-origin-when-cross-origin')
    response.headers.setdefault('Permissions-Policy', 'geolocation=(), microphone=(), camera=()')
    return response

@app.route('/')
def index():
    """Main Page - Dashboard """
    total_product = Product.query.count()
    total_stock = db.session.query(func.sum(Product.stock)).scalar() or 0
    critical_stock = Product.query.filter(Product.stock <= Product.min_stock).count()
    total_value = db.session.query(func.sum(Product.stock * Product.price)).scalar() or 0

    last_movements = StockMovement.query.order_by(StockMovement.date.desc()).limit(10).all()

    critical_products = Product.query.filter(Product.stock <= Product.min_stock).all()

    return render_template('index.html',
                       total_product = total_product,
                       total_stock = total_stock,
                       critical_stock = critical_stock,
                       total_value = total_value,
                       last_movements = last_movements,
                       critical_products = critical_products)

@app.route('/products')
def product_list():
    """Lists All Products"""
    page = request.args.get('page', 1, type=int)
    category_id = request.args.get('category', type=int)
    search = request.args.get('search', '')
    query = Product.query

    if category_id:
        query = query.filter_by(category_id=category_id)
    
    if search:
        query = query.filter(Product.name.like(f'%{search}%'))

    products = query.paginate(page=page, per_page=20, error_out=False)

    categories = Category.query.all()

    return render_template('product_list.html',
                           products=products,
                           categories=categories,
                           selected_category=category_id,
                           search=search)

@app.route('/products/export/excel')
def export_products_excel():
    """Export products to Excel"""
    category_id = request.args.get('category', type=int)
    search = request.args.get('search', '')
    
    query = Product.query
    if category_id:
        query = query.filter_by(category_id=category_id)
    if search:
        query = query.filter(Product.name.like(f'%{search}%'))
    
    products = query.all()
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from flask import send_file
    from datetime import datetime
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Headers
    headers = ['ID', 'Name', 'Barcode', 'Category', 'Price', 'Stock', 'Min Stock', 'Status', 'Total Value']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row, product in enumerate(products, 2):
        status = 'Critical' if product.critical_stock else 'Normal'
        total_value = product.stock * product.price
        
        data = [
            product.id,
            product.name,
            product.barcode or '',
            product.categorie.name,
            product.price,
            product.stock,
            product.min_stock,
            status,
            total_value
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            if col == 8 and value == 'Critical':  # Status column
                cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/product/delete/<int:id>', methods=['POST'])
def delete_product(id):
    """Delete a product"""
    from flask import jsonify
    
    product = Product.query.get_or_404(id)
    force_delete = request.form.get('force_delete') == 'true'
    
    # Check if product has stock movements
    movements = StockMovement.query.filter_by(product_id=id).all()
    if movements and not force_delete:
        # Return JSON response for AJAX handling
        return jsonify({
            'success': False,
            'has_movements': True,
            'movement_count': len(movements),
            'product_name': product.name,
            'message': f'Product "{product.name}" has {len(movements)} stock movements. Do you want to delete them too?'
        })
    
    product_name = product.name
    
    # If force delete, remove all stock movements first
    if force_delete and movements:
        for movement in movements:
            db.session.delete(movement)
        flash(f'Deleted {len(movements)} stock movements for product "{product_name}"', 'info')
    
    db.session.delete(product)
    db.session.commit()
    
    flash(f'Product "{product_name}" was deleted successfully!', 'success')
    
    # Return JSON response for AJAX
    return jsonify({
        'success': True,
        'message': f'Product "{product_name}" was deleted successfully!'
    })

@app.route('/product/add', methods=['GET', 'POST'])
def add_product():
    """Adds a new product"""
    form = ProductForm()

    form.category_id.choices = [(k.id, k.name) for k in Category.query.all()]

    if form.validate_on_submit():
        # Handle empty barcode
        barcode_value = form.barcode.data.strip() if form.barcode.data else None
        if barcode_value == '':
            barcode_value = None
            
        product = Product(
            name = form.name.data,
            barcode = barcode_value,
            price = form.price.data,
            stock = form.stock.data,
            min_stock = form.min_stock.data,
            category_id = form.category_id.data
        )
        db.session.add(product)
        db.session.commit()

        movement = StockMovement(
            product_id = product.id,
            type = 'inflow',
            amount=form.stock.data,
            previous_stock= 0,
            new_stock=form.stock.data,
            description='First stock inflow'
        )
        db.session.add(movement)
        db.session.commit()

        flash(f'{product.name} was saccsessfully added!', 'success')

        return redirect(url_for('product_list'))
    
    return render_template('product_add.html', form=form)

@app.route('/product/edit/<int:id>', methods=['GET', 'POST'])
def edit_product(id):
    """Edit Product"""
    product = Product.query.get_or_404(id)
    form = ProductForm(obj=product)
    form.category_id.choices = [(k.id, k.name) for k in Category.query.all()]

    if form.validate_on_submit():
        # Handle empty barcode
        barcode_value = form.barcode.data.strip() if form.barcode.data else None
        if barcode_value == '':
            barcode_value = None
            
        product.name = form.name.data
        product.barcode = barcode_value
        product.price = form.price.data
        product.min_stock = form.min_stock.data
        product.category_id = form.category_id.data

        db.session.commit()
        flash(f'{product.name} was updated!', 'success')
        return redirect(url_for('product_list'))
    return render_template('edit_product.html', form=form, product=product)

@app.route('/stock-movement', methods=['GET','POST'])
def stock_movement():
    """Stock inflow and outflow process"""

    form = StockMovementForm()

    form.product_id.choices = [(u.id, f'{u.id} (Stock: {u.stock})') for u in Product.query.all()]

    if form.validate_on_submit():
        product = Product.query.get(form.product_id.data)
        previous_stock = product.stock

        if form.type.data == 'inflow':
            product.stock += form.amount.data
        else:
            if product.stock >= form.amount.data:
                product.stock -= form.amount.data
            else:
                flash('Not enough stock!', 'danger')
                return render_template('stock_movement.html', form=form)
            
        movement = StockMovement(
            product_id = product.id,
            type= form.type.data,
            amount=form.amount.data,
            previous_stock=previous_stock,
            new_stock = product.stock,
            description = form.description.data
        )
        db.session.add(movement)
        db.session.commit()

        flash(f'Stock movement was saved! New Stock: {product.stock}', 'success')
        return redirect(url_for('index'))
    return render_template('stock_movement.html', form=form)

# CATEGORY MANAGEMENT ROUTES
@app.route('/categories')
def category_list():
    """Lists all categories"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    query = Category.query
    
    if search:
        query = query.filter(Category.name.like(f'%{search}%'))
    
    categories = query.paginate(page=page, per_page=20, error_out=False)
    
    return render_template('category_list.html', categories=categories, search=search)

@app.route('/category/add', methods=['GET', 'POST'])
def add_category():
    """Add a new category"""
    form = CategoryForm()
    
    if form.validate_on_submit():
        category = Category(
            name=form.name.data,
            description=form.description.data
        )
        db.session.add(category)
        db.session.commit()
        
        flash(f'Category "{category.name}" was successfully added!', 'success')
        return redirect(url_for('category_list'))
    
    return render_template('category_add.html', form=form)

@app.route('/category/edit/<int:id>', methods=['GET', 'POST'])
def edit_category(id):
    """Edit a category"""
    category = Category.query.get_or_404(id)
    form = CategoryForm(obj=category)
    
    if form.validate_on_submit():
        category.name = form.name.data
        category.description = form.description.data
        
        db.session.commit()
        flash(f'Category "{category.name}" was updated!', 'success')
        return redirect(url_for('category_list'))
    
    return render_template('category_edit.html', form=form, category=category)

@app.route('/category/delete/<int:id>', methods=['POST'])
def delete_category(id):
    """Delete a category"""
    category = Category.query.get_or_404(id)
    
    # Check if category has products
    if category.products:
        flash(f'Cannot delete category "{category.name}" because it has {len(category.products)} products!', 'danger')
        return redirect(url_for('category_list'))
    
    category_name = category.name
    db.session.delete(category)
    db.session.commit()
    
    flash(f'Category "{category_name}" was deleted!', 'success')
    return redirect(url_for('category_list'))

# ANALYTICS ROUTES
@app.route('/analytics')
def analytics():
    """Advanced Analytics Dashboard"""
    import pandas as pd
    from datetime import datetime, timedelta
    
    # Get all data
    products = Product.query.all()
    categories = Category.query.all()
    movements = StockMovement.query.all()
    
    # Stock Status Analysis
    total_products = len(products)
    critical_products = len([p for p in products if p.critical_stock])
    normal_products = total_products - critical_products
    
    stock_status_data = {
        'labels': ['Normal Stock', 'Critical Stock'],
        'data': [normal_products, critical_products],
        'colors': ['#28a745', '#dc3545']
    }
    
    # Category Distribution
    category_data = {}
    for category in categories:
        category_data[category.name] = len(category.products)
    
    category_chart_data = {
        'labels': list(category_data.keys()),
        'data': list(category_data.values()),
        'colors': ['#007bff', '#28a745', '#ffc107', '#dc3545', '#6f42c1', '#fd7e14', '#20c997', '#e83e8c']
    }
    
    # Stock Value by Category
    category_values = {}
    for category in categories:
        total_value = sum(p.stock * p.price for p in category.products)
        category_values[category.name] = round(total_value, 2)
    
    value_chart_data = {
        'labels': list(category_values.keys()),
        'data': list(category_values.values()),
        'colors': ['#007bff', '#28a745', '#ffc107', '#dc3545', '#6f42c1', '#fd7e14', '#20c997', '#e83e8c']
    }
    
    # Stock Movement Trends (Last 30 days)
    thirty_days_ago = datetime.now() - timedelta(days=30)
    recent_movements = [m for m in movements if m.date >= thirty_days_ago]
    
    # Group movements by date
    movement_by_date = {}
    for movement in recent_movements:
        date_str = movement.date.strftime('%Y-%m-%d')
        if date_str not in movement_by_date:
            movement_by_date[date_str] = {'inflow': 0, 'outflow': 0}
        movement_by_date[date_str][movement.type] += movement.amount
    
    # Sort by date
    sorted_dates = sorted(movement_by_date.keys())
    
    trend_data = {
        'labels': sorted_dates,
        'inflow': [movement_by_date[date]['inflow'] for date in sorted_dates],
        'outflow': [movement_by_date[date]['outflow'] for date in sorted_dates]
    }
    
    # Top Products by Stock Value
    product_values = [(p.name, p.stock * p.price) for p in products]
    product_values.sort(key=lambda x: x[1], reverse=True)
    top_products = product_values[:10]
    
    top_products_data = {
        'labels': [p[0] for p in top_products],
        'data': [round(p[1], 2) for p in top_products],
        'colors': ['#007bff'] * len(top_products)
    }
    
    # Calculate summary statistics
    total_stock_value = sum(p.stock * p.price for p in products)
    total_inflow = sum(m.amount for m in recent_movements if m.type == 'inflow')
    total_outflow = sum(m.amount for m in recent_movements if m.type == 'outflow')
    
    # Low stock alerts
    low_stock_products = [p for p in products if p.stock <= p.min_stock * 1.2]  # 20% above minimum
    
    analytics_data = {
        'stock_status': stock_status_data,
        'category_distribution': category_chart_data,
        'category_values': value_chart_data,
        'stock_trends': trend_data,
        'top_products': top_products_data,
        'summary': {
            'total_products': total_products,
            'critical_products': critical_products,
            'total_categories': len(categories),
            'total_stock_value': round(total_stock_value, 2),
            'total_inflow': total_inflow,
            'total_outflow': total_outflow,
            'net_movement': total_inflow - total_outflow,
            'low_stock_count': len(low_stock_products)
        },
        'low_stock_products': low_stock_products
    }
    
    return render_template('analytics.html', data=analytics_data)

# REPORT ROUTES
@app.route('/reports')
def reports():
    """Reports page"""
    form = ReportForm()
    form.category_id.choices = [('', 'All Categories')] + [(c.id, c.name) for c in Category.query.all()]
    
    # Default date range (last 30 days)
    from datetime import date, timedelta
    form.starting_date.data = date.today() - timedelta(days=30)
    form.ending_date.data = date.today()
    
    return render_template('reports.html', form=form)

@app.route('/reports/generate', methods=['POST'])
def generate_report():
    """Generate report based on filters"""
    form = ReportForm()
    form.category_id.choices = [('', 'All Categories')] + [(c.id, c.name) for c in Category.query.all()]
    
    if form.validate_on_submit():
        # Filter products
        query = Product.query
        category_id = None
        if form.category_id.data and form.category_id.data != '':
            try:
                category_id = int(form.category_id.data)
            except (ValueError, TypeError):
                category_id = None
        
        if category_id:
            query = query.filter_by(category_id=category_id)
        
        products = query.all()
        
        # Filter stock movements
        movement_query = StockMovement.query.filter(
            StockMovement.date >= form.starting_date.data,
            StockMovement.date <= form.ending_date.data
        )
        
        if category_id:
            movement_query = movement_query.join(Product).filter(Product.category_id == category_id)
        
        movements = movement_query.order_by(StockMovement.date.desc()).all()
        
        # Calculate statistics
        total_inflow = sum(m.amount for m in movements if m.type == 'inflow')
        total_outflow = sum(m.amount for m in movements if m.type == 'outflow')
        
        report_data = {
            'products': products,
            'movements': movements,
            'total_inflow': total_inflow,
            'total_outflow': total_outflow,
            'net_movement': total_inflow - total_outflow,
            'date_range': {
                'start': form.starting_date.data,
                'end': form.ending_date.data
            },
            'category_filter': Category.query.get(category_id) if category_id else None
        }
        
        return render_template('report_results.html', data=report_data, form=form)
    
    return render_template('reports.html', form=form)

@app.route('/reports/export/excel')
def export_excel():
    """Export report to Excel"""
    # Get filter parameters
    category_id_param = request.args.get('category_id', '')
    category_id = None
    if category_id_param not in (None, ''):
        try:
            category_id = int(category_id_param)
        except (TypeError, ValueError):
            abort(400, description='Invalid category identifier')

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    start = end = None
    if start_date and end_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            abort(400, description='Invalid date range supplied')
    
    # Create Excel file
    wb = Workbook()
    
    # Products sheet
    ws_products = wb.active
    ws_products.title = "Products"
    ws_products.append(['ID', 'Name', 'Barcode', 'Category', 'Price', 'Stock', 'Min Stock', 'Status', 'Total Value'])
    
    query = Product.query
    if category_id:
        query = query.filter_by(category_id=category_id)
    
    for product in query.all():
        status = 'Critical' if product.critical_stock else 'Normal'
        total_value = product.stock * product.price
        ws_products.append([
            product.id,
            product.name,
            product.barcode or '',
            product.categorie.name,
            product.price,
            product.stock,
            product.min_stock,
            status,
            total_value
        ])
    
    # Stock Movements sheet
    ws_movements = wb.create_sheet("Stock Movements")
    ws_movements.append(['Date', 'Product', 'Type', 'Amount', 'Previous Stock', 'New Stock', 'Description'])
    
    movement_query = StockMovement.query
    if start and end:
        movement_query = movement_query.filter(
            StockMovement.date >= start,
            StockMovement.date <= end
        )
    
    if category_id:
        movement_query = movement_query.join(Product).filter(Product.category_id == category_id)
    
    for movement in movement_query.order_by(StockMovement.date.desc()).all():
        ws_movements.append([
            movement.date.strftime('%Y-%m-%d %H:%M'),
            movement.product.name,
            movement.type.title(),
            movement.amount,
            movement.previous_stock,
            movement.new_stock,
            movement.description or ''
        ])
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"stock_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/reports/export/pdf')
def export_pdf():
    """Export detailed report to PDF"""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    
    # Get filter parameters
    category_id_param = request.args.get('category_id', '')
    category_id = None
    if category_id_param not in (None, ''):
        try:
            category_id = int(category_id_param)
        except (TypeError, ValueError):
            abort(400, description='Invalid category identifier')

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    start = end = None
    if start_date and end_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d').date()
            end = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            abort(400, description='Invalid date range supplied')
    
    # Create PDF
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    story = []
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=26,
        spaceAfter=30,
        alignment=1,
        textColor=colors.darkblue
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=15,
        spaceBefore=20,
        textColor=colors.darkblue,
        borderWidth=1,
        borderColor=colors.darkblue,
        borderPadding=5,
        backColor=colors.lightblue
    )
    
    # Header
    story.append(Paragraph("📊 DETAILED STOCK REPORT", title_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %H:%M:%S')}", styles['Normal']))
    
    # Report parameters
    if category_id or start_date or end_date:
        story.append(Spacer(1, 15))
        story.append(Paragraph("🔍 REPORT FILTERS", heading_style))
        
        if category_id:
            category = Category.query.get(category_id)
            story.append(Paragraph(f"📁 Category: <b>{category.name}</b>", styles['Normal']))
        else:
            story.append(Paragraph("📁 Category: <b>All Categories</b>", styles['Normal']))
            
        if start and end:
            story.append(Paragraph(f"📅 Date Range: <b>{start.strftime('%Y-%m-%d')} to {end.strftime('%Y-%m-%d')}</b>", styles['Normal']))
        else:
            story.append(Paragraph("📅 Date Range: <b>All Time</b>", styles['Normal']))
    
    story.append(Spacer(1, 30))
    
    # Products section
    story.append(Paragraph("📦 PRODUCTS SUMMARY", heading_style))
    
    product_data = [['Product Name', 'Category', 'Stock', 'Min Stock', 'Status', 'Unit Value', 'Total Value']]
    
    query = Product.query
    if category_id:
        query = query.filter_by(category_id=category_id)
    
    total_products = 0
    total_value = 0
    critical_count = 0
    
    for product in query.all():
        status = '⚠️ Critical' if product.critical_stock else '✅ Normal'
        if product.critical_stock:
            critical_count += 1
        unit_value = f"${product.price:.2f}"
        product_total_value = product.stock * product.price
        total_value += product_total_value
        total_products += 1
        
        product_data.append([
            product.name,
            product.categorie.name,
            str(product.stock),
            str(product.min_stock),
            status,
            unit_value,
            f"${product_total_value:.2f}"
        ])
    
    # Summary row
    product_data.append([
        f'TOTAL ({total_products} products)',
        f'{critical_count} Critical Items',
        '-',
        '-',
        f'{critical_count}/{total_products}',
        '-',
        f'${total_value:,.2f}'
    ])
    
    if len(product_data) > 2:  # More than just header and summary
        product_table = Table(product_data, colWidths=[2*inch, 1.3*inch, 0.7*inch, 0.7*inch, 1*inch, 0.8*inch, 1*inch])
        product_table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows
            ('BACKGROUND', (0, 1), (-1, -2), colors.lightgrey),
            ('GRID', (0, 0), (-1, -2), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -2), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey]),
            
            # Summary row
            ('BACKGROUND', (0, -1), (-1, -1), colors.darkgreen),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 10)
        ]))
        
        story.append(product_table)
    else:
        story.append(Paragraph("No products found matching the criteria.", styles['Normal']))
    
    # Stock movements section
    if start and end:
        story.append(Spacer(1, 30))
        story.append(Paragraph("📋 STOCK MOVEMENTS", heading_style))
        
        movement_query = StockMovement.query.filter(
            StockMovement.date >= start,
            StockMovement.date <= end
        )
        
        if category_id:
            movement_query = movement_query.join(Product).filter(Product.category_id == category_id)
        
        movements = movement_query.order_by(StockMovement.date.desc()).all()
        
        if movements:
            movement_data = [['Date', 'Product', 'Type', 'Amount', 'Previous', 'New Stock', 'Description']]
            
            total_inflow = 0
            total_outflow = 0
            
            for movement in movements:
                if movement.type == 'inflow':
                    total_inflow += movement.amount
                    type_display = '📈 Inflow'
                else:
                    total_outflow += movement.amount
                    type_display = '📉 Outflow'
                
                movement_data.append([
                    movement.date.strftime('%Y-%m-%d'),
                    movement.product.name,
                    type_display,
                    str(movement.amount),
                    str(movement.previous_stock),
                    str(movement.new_stock),
                    movement.description[:25] + '...' if movement.description and len(movement.description) > 25 else (movement.description or 'N/A')
                ])
            
            # Summary row for movements
            movement_data.append([
                f'SUMMARY ({len(movements)} movements)',
                f'Inflow: {total_inflow}',
                f'Outflow: {total_outflow}',
                f'Net: {total_inflow - total_outflow}',
                '-',
                '-',
                f"Period: {start.strftime('%Y-%m-%d')} to {end.strftime('%Y-%m-%d')}"
            ])
            
            movement_table = Table(movement_data, colWidths=[1*inch, 1.5*inch, 0.8*inch, 0.6*inch, 0.6*inch, 0.6*inch, 1.4*inch])
            movement_table.setStyle(TableStyle([
                # Header
                ('BACKGROUND', (0, 0), (-1, 0), colors.purple),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Data rows
                ('BACKGROUND', (0, 1), (-1, -2), colors.lavender),
                ('GRID', (0, 0), (-1, -2), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -2), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lavender]),
                
                # Summary row
                ('BACKGROUND', (0, -1), (-1, -1), colors.darkred),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 9)
            ]))
            
            story.append(movement_table)
        else:
            story.append(Paragraph("No stock movements found in the specified date range.", styles['Normal']))
    
    # Footer
    story.append(Spacer(1, 30))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        alignment=1,
        textColor=colors.grey
    )
    story.append(Paragraph("Generated by Stock Tracking System | Detailed Report", footer_style))
    
    # Build PDF
    doc.build(story)
    output.seek(0)
    
    filename = f"stock_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/pdf'
    )

@app.route('/dashboard/export/pdf')
def export_dashboard_pdf():
    """Export dashboard summary to PDF"""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from datetime import datetime, timedelta
    
    # Get dashboard data
    products = Product.query.all()
    categories = Category.query.all()
    movements = StockMovement.query.filter(
        StockMovement.date >= datetime.now() - timedelta(days=30)
    ).all()
    
    # Calculate statistics
    total_products = len(products)
    critical_products = len([p for p in products if p.critical_stock])
    total_value = sum(p.stock * p.price for p in products)
    total_inflow = sum(m.amount for m in movements if m.type == 'inflow')
    total_outflow = sum(m.amount for m in movements if m.type == 'outflow')
    
    # Create PDF
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    story = []
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=26,
        spaceAfter=30,
        alignment=1,
        textColor=colors.darkblue
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=15,
        spaceBefore=20,
        textColor=colors.darkblue,
        borderWidth=1,
        borderColor=colors.darkblue,
        borderPadding=5,
        backColor=colors.lightblue
    )
    
    # Header
    story.append(Paragraph("📊 STOCK TRACKING DASHBOARD REPORT", title_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %H:%M:%S')}", styles['Normal']))
    story.append(Spacer(1, 30))
    
    # Summary Statistics Section
    story.append(Paragraph("📈 SUMMARY STATISTICS", heading_style))
    
    summary_data = [
        ['Metric', 'Value', 'Status'],
        ['Total Products', str(total_products), '✅ Active'],
        ['Critical Stock Items', str(critical_products), '⚠️ Attention Needed' if critical_products > 0 else '✅ All Good'],
        ['Total Categories', str(len(categories)), '📁 Organized'],
        ['Total Stock Value', f'${total_value:,.2f}', '💰 Asset Value'],
        ['Inflow (30 days)', str(total_inflow), '📈 Received'],
        ['Outflow (30 days)', str(total_outflow), '📉 Consumed'],
        ['Net Movement', str(total_inflow - total_outflow), '⚖️ Balance']
    ]
    
    summary_table = Table(summary_data, colWidths=[2.5*inch, 1.5*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Critical Stock Items
    if critical_products > 0:
        story.append(Paragraph("⚠️ CRITICAL STOCK ALERTS", heading_style))
        
        critical_data = [['Product Name', 'Category', 'Current Stock', 'Min Required', 'Action Needed']]
        for product in products:
            if product.critical_stock:
                critical_data.append([
                    product.name,
                    product.categorie.name,
                    str(product.stock),
                    str(product.min_stock),
                    'RESTOCK IMMEDIATELY'
                ])
        
        critical_table = Table(critical_data, colWidths=[2*inch, 1.5*inch, 1*inch, 1*inch, 1.5*inch])
        critical_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.red),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.mistyrose),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9)
        ]))
        
        story.append(critical_table)
        story.append(Spacer(1, 20))
    
    # Category Breakdown
    story.append(Paragraph("📁 CATEGORY BREAKDOWN", heading_style))
    
    category_data = [['Category', 'Products Count', 'Total Value', 'Avg Value per Product']]
    for category in categories:
        products_count = len(category.products)
        if products_count > 0:
            category_value = sum(p.stock * p.price for p in category.products)
            avg_value = category_value / products_count
            category_data.append([
                category.name,
                str(products_count),
                f'${category_value:,.2f}',
                f'${avg_value:,.2f}'
            ])
    
    category_table = Table(category_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    category_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgreen])
    ]))
    
    story.append(category_table)
    story.append(Spacer(1, 20))
    
    # Recent Stock Movements (Last 10)
    story.append(Paragraph("📋 RECENT STOCK MOVEMENTS", heading_style))
    
    recent_movements = movements[-10:] if len(movements) > 10 else movements
    movement_data = [['Date', 'Product', 'Type', 'Amount', 'Description']]
    
    for movement in reversed(recent_movements):
        movement_data.append([
            movement.date.strftime('%Y-%m-%d'),
            movement.product.name,
            f"{'📈' if movement.type == 'inflow' else '📉'} {movement.type.title()}",
            str(movement.amount),
            movement.description[:30] + '...' if movement.description and len(movement.description) > 30 else (movement.description or 'N/A')
        ])
    
    if len(movement_data) > 1:
        movement_table = Table(movement_data, colWidths=[1.2*inch, 2*inch, 1.3*inch, 0.8*inch, 2.2*inch])
        movement_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.purple),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lavender),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lavender])
        ]))
        
        story.append(movement_table)
    else:
        story.append(Paragraph("No recent stock movements found.", styles['Normal']))
    
    # Footer
    story.append(Spacer(1, 30))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        alignment=1,
        textColor=colors.grey
    )
    story.append(Paragraph("Generated by Stock Tracking System | Dashboard Report", footer_style))
    
    # Build PDF
    doc.build(story)
    output.seek(0)
    
    filename = f"dashboard_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/pdf'
    )

if __name__ == '__main__':
    app.run(debug=True, host='127.0.0.1', port=5000)
    